import sys
import os
from math import pi
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from PyQt5 import QtWidgets as qtw
from PyQt5 import QtCore as qtc
from PyQt5 import QtGui as qtg
from PyQt5 import uic
from datetime import datetime
import time

MyWindow, base_class = uic.loadUiType('./gui.ui')
class MainWindow(base_class):
    def __init__(self, *args,**kwargs):
        super().__init__(*args,**kwargs)
        self.ui = MyWindow()
        self.ui.setupUi(self)
        
        self.initialize()

        
    def initialize(self):
        self.setWindowTitle('ADT to Excel')

        self.ui.new_done_tick.setHidden(True)
        self.ui.existing_done_tick.setHidden(True)
        self.timer=qtc.QTimer()
        self.timer.timeout.connect(lambda: self.ui.new_done_tick.setHidden(True))
        self.timer.timeout.connect(lambda: self.ui.existing_done_tick.setHidden(True))

        #BUTTONS
        self.ui.select_txt_btn.clicked.connect(lambda: self.load_file('Txt','txt'))
        self.ui.select_existing_excel_btn.clicked.connect(lambda: self.load_file('Excel','xlsx'))
        self.ui.write_new_btn.clicked.connect(self.write_new_button_pressed)
        self.ui.write_existing_btn.clicked.connect(self.write_existing_btn_pressed)


    def load_file(self, format, extenion):
        if format == 'Excel':
            tup_file_path = qtw.QFileDialog.getOpenFileName(self, 'Open File',filter=f"{format} (*.{extenion})")
            if not tup_file_path[0]:
                
                return None
            file_path = tup_file_path[0]    
            self.ui.selected_xlsx_lbl.setText(f'Excel выбран:\n{os.path.basename(file_path)}')
            self.excel_file_path = file_path
        else:
            tup_files_paths = qtw.QFileDialog.getOpenFileNames(self, 'Open File',filter=f"{format} (*.{extenion})")
            if not tup_files_paths[0]:
                return None
            files_paths = tup_files_paths[0]
            self.ui.selected_txt_lbl.setText(f'Выбрано:\n{len(files_paths)} txt')
            self.txt_files_paths = files_paths
            for widget in self.ui.RightMenu.findChildren((qtw.QLabel,qtw.QPushButton,qtw.QLineEdit)):
                widget.setEnabled(True)

    def write_new_button_pressed(self):
        if not hasattr(self, 'txt_files_paths'):
            return None
        logs = self.compose_logs_from_files()

        wb = Workbook()
        ws = wb.active
        alignment = Alignment(horizontal='center')
        for row in ws.iter_rows(min_row = 1, max_row = len(logs), min_col = 1, max_col = len(logs[0])):
            for cell in row:
                cell.value = logs[cell.row-1][cell.column-1]
                cell.alignment = alignment

        self.set_styles(ws,len(logs[0]))
        if self.ui.file_name_lineEdit.text() == 'logs timestamp':
            wb.save(f'logs {datetime.now().strftime("%d.%m.%Y %H-%M-%S")}.xlsx')
        else:
            wb.save(f'{self.ui.file_name_lineEdit.text()}.xlsx')
        self.ui.new_done_tick.setHidden(False)
        self.timer.start(1000)
        
    def write_existing_btn_pressed(self):
        #TODO catch opened excel exception
        if not hasattr(self, 'excel_file_path'):
            return None

        logs = self.compose_logs_from_files()
        logs.pop(0)
        wb = load_workbook(self.excel_file_path)
        sh = wb.active
        
        alignment = Alignment(horizontal='center')
        end_of_old_logs = sh.max_row
        for row in sh.iter_rows(min_row = end_of_old_logs+1, max_row = len(logs) + end_of_old_logs, min_col = 1, max_col = len(logs[0])):
            for cell in row:
                cell.value = logs[cell.row-1-end_of_old_logs][cell.column-1]
                cell.alignment = alignment
        wb.save(os.path.basename(self.excel_file_path))
        self.ui.existing_done_tick.setHidden(False)
        self.timer.start(1000)

    def set_styles(self, ws, length):
        bold = Font(color="00000000", bold  = True)
        
        for row in ws.iter_rows(min_row = 1, max_row = 1, min_col=1, max_col = length):
            for cell in row:
                cell.font = bold
                ws.column_dimensions[get_column_letter(cell.column)].width = 12

    def compose_logs_from_files(self):
        logs = [['Date', 'Origin time', 'Lat', 'Lon', 'AzMajor', 'Rminor', 'Rmajor', 'S']]
        events = []    # 1event = text in 1 file 
        for txt_file_path in self.txt_files_paths:
            with open(txt_file_path, 'r', encoding='ANSI') as file:
                text = file.read()
                events.append(text)
        
        for event in events:
            event_log = []
            event = event.strip()
            lines = event.split('\n')
            del lines[1]
            line1, line2 = lines
            
            dating = line1[line1.find('T0=')+3:event.find('Err=')-1]
            fulldate = datetime.strptime(dating, "%d.%m.%Y   %H.%M:%S.%f")
            date = fulldate.strftime("%d.%m.%Y")
            time = fulldate.strftime('%H:%M:%S.%f')[:-5]
            
            lat = float(line1[line1.find('FI=')+3:line1.find('LD=')-1])
            lon = float(line1[line1.find('LD=')+3:line1.find('T0=')-1])
            
            azimuth = line2[line2.find('Азимут : '):line2.find('большой')-1].split(' : ')[1]
            Rmajor = line2[line2.find('большой радиус : '):line2.find('малый радиус')].split(' : ')[1]
            Rminor = line2[line2.find('малый радиус : '):].split(' : ')[1] 
            S = float(Rminor) * float(Rmajor) * pi

            event_log = [date,time,round(lat,3),round(lon,3), float(azimuth), float(Rminor), float(Rmajor), round(S,2)]
            logs.append(event_log)
        return logs


if __name__ == '__main__':
    app = qtw.QApplication(sys.argv)
    mw = MainWindow()
    mw.show()
    sys.exit(app.exec_())